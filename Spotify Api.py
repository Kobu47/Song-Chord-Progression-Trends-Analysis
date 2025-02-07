import pandas as pd
import requests
import openpyxl
import base64
from openpyxl import load_workbook


redirect_url = "http://localhost:3000/callback"
base_url = "https://accounts.spotify.com/api/"

#Credentials
client_id = "3d8fa969aeef49b6ad3ee859705290b7"
client_secret = "28d83493c0ab4997847bcc5f99efa751"

#Encoded Credentials using Base64
auth_string = f"{client_id}:{client_secret}"
auth_bytes = auth_string.encode("utf-8")
auth_base64 = base64.b64encode(auth_bytes).decode("utf-8")

data = {
    "grant_type": "client_credentials",
}
auth_headers = {
    "Authorization": f"Basic {auth_base64}",
    "Content-Type" : "application/x-www-form-urlencoded"
}

#Requests access token which expires in 1 hour.
def token_request():
    url = f"{base_url}token"
    response = requests.post(url, headers=auth_headers ,data=data)

    if response.status_code == 200:
        token = response.json().get("access_token")
        return token
    else:
        print(f"failed to receive token {response.status_code}")

#Requests specifically the popularity number of a song based on it's name.
def song_search(song_name,token):
    url = f"https://api.spotify.com/v1/search?q={song_name}&type=track&limit=1"
    headers = {
    "Authorization": f"Bearer {token}"
}
    response = requests.get(url , headers=headers)

    if response.status_code == 200:
        results = response.json()
        song_result = results["tracks"]["items"][0]
        print(song_result["popularity"], song_result["album"]["release_date"]
        )
        return{
            "artist_id" : song_result["artists"][0]["id"],
            "popularity" : song_result["popularity"],
            "date" : song_result["album"]["release_date"]
        }
    else:
        print(f"There was no results {response.status_code}")

#Appends the data to an already existing Excel file.
def append_to_excel_file(file,dataframe,sheet):
    song_data = pd.DataFrame([dataframe])
    book = load_workbook(file)
    file_data = pd.read_excel(file,sheet_name=sheet)
    combined_data = pd.concat([file_data,song_data])
    with  pd.ExcelWriter(file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        combined_data.to_excel(writer, sheet_name = sheet, index = False)
    print("Data updated successfully")

#Saves the data into an Excel file.
def data_to_excel(dataframe):
    dataframe.to_excel("Spotify_data.xlsx", index = False)
    print("Data was saved successfully in an Excel file")

#Gets the genre of the artist
def artist_genre(id,token):
    url = f"https://api.spotify.com/v1/artists/{id}"
    headers = {
        "Authorization": f"Bearer {token}"
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        artist = response.json()
        return artist["genres"]
    else:
        print(f"Failed to get genres{response.status_code}")


token = token_request()
popularity = []
genre = []
release_date = []
book = load_workbook("Alpha_data.xlsx")
existing_data = pd.read_excel("Alpha_data.xlsx",sheet_name="Songs")
for song_name in existing_data["song"]:
    result_dict = song_search(song_name,token)
    popularity.append(result_dict.get("popularity"))
    genres_response = artist_genre((result_dict.get("artist_id")),token)
    genre.append(genres_response)
    release_date.append(result_dict.get("date"))

existing_data["Popularity"] = popularity
existing_data["Genre"] = genre
existing_data["Release Date"] = release_date
data_to_excel(existing_data)